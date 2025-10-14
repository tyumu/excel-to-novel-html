"""
Excelデータを一気にHTMLに変換するシンプル版

このスクリプトは:
1. Excelから全データを抽出
2. 見やすく整形して表示
3. あなたがコピペしてAIに送信
4. AIの結果をコピペして入力
5. 自動でHTMLを生成

使い方: python simple_converter.py
"""

import re
from pathlib import Path

# デフォルトのExcelファイル名
DEFAULT_EXCEL_FILE = 'main_0_暗黒時代・上.xlsx'

# openpyxlはExcel抽出時のみ必要
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 設定ファイルを読み込み（存在する場合）
try:
    from config import DOCTOR_NAME, BRANCH_MODE, BRANCH_DISPLAY
except ImportError:
    # デフォルト設定
    DOCTOR_NAME = "ドクター"
    BRANCH_MODE = "include_all"
    BRANCH_DISPLAY = {"show_options": True, "options_format": "inline"}

def extract_title_from_filename(filename):
    """ファイル名からタイトルを抽出
    例: main_0_暗黒時代・上.xlsx → main_0_暗黒時代・上
    """
    match = re.match(r'(main_\d+_.+)\.xlsx', filename)
    if match:
        return match.group(1)
    # マッチしない場合は拡張子を除去
    return filename.replace('.xlsx', '')

def extract_all_dialogues(excel_file=None):
    """Excelから全てのシートの会話を抽出（話者情報込み）"""
    if not HAS_OPENPYXL:
        print("エラー: openpyxlがインストールされていません")
        print("インストール方法: pip install openpyxl")
        return None
    
    if excel_file is None:
        excel_file = DEFAULT_EXCEL_FILE
    
    wb = openpyxl.load_workbook(excel_file)
    
    print("=" * 80)
    print(f"エクセルから会話データを抽出中: {excel_file}")
    print("=" * 80)
    
    all_text = []
    decision_count = 0
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        all_text.append(f"\n\n{'=' * 60}")
        all_text.append(f"【シーン: {sheet_name}】")
        all_text.append(f"{'=' * 60}\n")
        
        in_decision = False
        in_branch = False
        current_options = []
        choices_already_displayed = False  # 選択肢を一度表示したかどうか
        
        for row in range(1, sheet.max_row + 1):
            col2 = sheet.cell(row=row, column=2).value  # 話者
            col3 = sheet.cell(row=row, column=3).value  # セリフ/内容
            
            if not col3:
                continue
                
            # 分岐システムの処理
            if col2 == '--Decision--':
                in_decision = True
                decision_count += 1
                if BRANCH_DISPLAY["show_options"] and BRANCH_DISPLAY["options_format"] == "separate_page":
                    all_text.append("\n【選択肢】")
                current_options = []
                choices_already_displayed = False  # リセット
                continue
                
            elif col2 == '--Decision End--':
                in_decision = False
                if current_options and BRANCH_MODE == "include_all" and not choices_already_displayed:
                    # 全ての選択肢をインラインで表示（最初の1回だけ）
                    if BRANCH_DISPLAY["show_options"] and BRANCH_DISPLAY["options_format"] == "inline":
                        all_text.append("\n【ドクターの選択肢】")
                        for i, opt in enumerate(current_options, 1):
                            all_text.append(f"  選択肢{i}: {opt}")
                        all_text.append("")
                        choices_already_displayed = True
                continue
                
            elif col2 and col2.startswith('Option_'):
                # 選択肢
                if in_decision:
                    current_options.append(col3.strip())
                continue
                
            elif col2 == '--Branch--':
                # 分岐点
                branch_info = col3.strip() if col3 else ""
                if BRANCH_MODE == "include_all":
                    # 分岐マーカーの後に選択肢番号を表示
                    all_text.append(f"\n【分岐: {branch_info}】")
                    
                    # >Options_X または >Options_X&Y&Z から番号を抽出
                    import re as re_module
                    # >Options_1&2&3 のような複数選択肢にも対応
                    match = re_module.search(r'>Options_([0-9&]+)', branch_info)
                    if match and current_options:
                        option_nums_str = match.group(1)
                        # &で分割して複数の番号を取得
                        option_nums = [int(n) for n in option_nums_str.split('&')]
                        # 各選択肢を表示
                        for option_num in option_nums:
                            if option_num <= len(current_options):
                                all_text.append(f"  選択肢{option_num}: {current_options[option_num-1]}")
                continue
            
            # 通常の処理
            if col2 == '--image--':
                all_text.append(f"[画像]: {col3}")
            elif col2 == '--background--':
                all_text.append(f"[背景]: {col3}")
            elif col2 in ['----', '--imagetween--']:
                continue
            elif col2 and col2 not in ['Option_1', 'Option_2', 'Option_3']:
                # {@nickname}を置き換え
                text = col3.strip().replace('{@nickname}', DOCTOR_NAME)
                all_text.append(f"【{col2}】{text}")
            else:
                # 話者情報がない場合
                text = col3.strip().replace('{@nickname}', DOCTOR_NAME)
                all_text.append(text)
    
    result = '\n'.join(all_text)
    print(f"\n検出した分岐数: {decision_count}")
    print(f"ドクターの名前: {DOCTOR_NAME}")
    print(f"分岐モード: {BRANCH_MODE}")
    
    return result

def save_to_file(content, filename):
    """ファイルに保存"""
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    
    filepath = output_dir / filename
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)
    
    return filepath

def create_html(novel_text, output_file='generated_novel.html', title='小説'):
    """小説テキストからHTMLを生成"""
    
    # HTMLテンプレート
    html_template = """<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    
    <style>
        /* 基本設定 */
        html, body {
            margin: 0;
            padding: 0;
            width: 100%;
            height: 100%;
            overflow: hidden; /* 意図しないスクロールを禁止 */
        }
        body {
            display: flex;
            /* ★修正点: 右から左へページが進むように */
            flex-direction: row-reverse; 
            
            overflow-x: scroll;
            overflow-y: hidden;
            -webkit-overflow-scrolling: touch;
            
            background-color: #FDFCF7;
            
            /* ★修正点: ページ間隔を最小限に */
            gap: 0;
        }

        /* すべての「ページ」に共通する設定 */
        .page {
            writing-mode: vertical-rl;
            text-orientation: mixed;
            
            /* ★修正点: 下のタブバーを考慮した余白 */
            width: calc(100vw - 10em);
            min-height: 50vh; /* 最小高さを設定 */
            max-height: calc(100vh - 4em); /* 最大高さは画面サイズ */
            padding-top: 2em;
            padding-bottom: 5em;
            padding-left: 2em;
            padding-right: 2em;
            
            flex-shrink: 0;
            
            font-family: 'Hiragino Mincho ProN', 'Yu Mincho', 'MS Mincho', serif;
            font-size: 16px;
            color: #333;
            line-height: 2.4;
            letter-spacing: 0.08em;
            
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            
            margin: 0;
            
        }

        /* テキストページの文字揃え */
        .page.text {
            align-items: stretch;
            text-align: justify;
            width: auto; /* テキストページは内容に合わせて幅を調整 */
            height: auto; /* テキストページは内容に合わせて高さ調整 */
            padding-left: 0.2em; /* テキストページのpadding調整 */
            padding-right: 0.2em;
        }
        
        .page:has(img.illustration) {
            width: calc(100vw - 2em); /* 画像ページの幅を広げる */
            padding: 1em;
            width: auto; /* テキストページは内容に合わせて幅を調整 */
            height: auto; /* テキストページは内容に合わせて高さ調整 */
            padding-left: 0.2em; /* テキストページのpadding調整 */
            padding-right: 0.2em;
        }
        
        /* 分岐表示用のスタイル */
        .page.branch {
            background-color: #F5F0E8;
            border-left: 3px solid #8B7355;
        }
        
        .branch-marker {
            font-weight: bold;
            color: #8B4513;
            margin-bottom: 1em;
        }
        
        .choice-header {
            font-weight: bold;
            color: #2C5F2D;
            margin: 1em 0 0.5em 0;
            padding: 0.5em;
            background-color: #E8F5E9;
            border-radius: 4px;
        }
        
        .choice-text {
            font-weight: bold;
            color: #1565C0;
            font-size: 1.1em;
            margin: 0.5em 0;
            padding: 0.3em;
            background-color: #E3F2FD;
            border-left: 3px solid #1976D2;
        }
        
        /* 見出しの設定 */
        h1, h2 {
            text-align: center;
            margin: 0;
        }
        h1 { font-size: 2em; }
        h2 { font-size: 1.5em; }
        h3 {
            font-size: 1.2em;
            color: #8B4513;
            text-align: center;
            margin: 0;
        }

        /* 話者名の設定 */
        .speaker {
            font-weight: bold;
            color: #2C5F2D;
            font-size: 0.95em;
            display: inline-block;
        }

        /* 画像の設定 */
        img.illustration {
             max-width: 100%;
             max-height: 95vh;
             width: auto;
             height: auto;
             object-fit: contain;
             display: block;
             margin: auto;
        }

        /* 縦中横の設定（数字や短い英語を横向きに表示） */
        .tcy {
            text-combine-upright: all;
            -webkit-text-combine: horizontal;
            -ms-text-combine-horizontal: all;
        }

        p {
             margin-top: 0;
             margin-bottom: 0;
        }
    </style>
</head>

<body>
    <div class="page">
        <h1>{title}</h1>
    </div>

{pages}
</body>
</html>"""
    
    # テキストをページに分割
    pages_html = []
    paragraphs = novel_text.split('\n\n')
    
    # 選択肢セクションを次の分岐セクションと結合する前処理
    merged_paragraphs = []
    i = 0
    while i < len(paragraphs):
        para = paragraphs[i].strip()
        
        # 選択肢セクションを検出
        if '【ドクターの選択肢】' in para or (para and para.split('\n')[0].strip().startswith('選択肢')):
            # 次の段落以降で分岐セクションを探して結合
            combined = [para]
            j = i + 1
            while j < len(paragraphs):
                next_para = paragraphs[j].strip()
                combined.append(next_para)
                # 分岐セクションの終わり（次の選択肢または通常テキスト）を検出
                if '【ドクターの選択肢】' in next_para or (next_para and not next_para.startswith('【分岐:') and '【分岐:' not in next_para):
                    # 分岐セクションが終わったので、次の選択肢の前で止まる
                    if '【ドクターの選択肢】' in next_para:
                        # 次の選択肢セクションなので含めない
                        combined.pop()
                        j -= 1
                    break
                j += 1
            merged_paragraphs.append('\n\n'.join(combined))
            i = j + 1
        else:
            merged_paragraphs.append(para)
            i += 1
    
    for para in merged_paragraphs:
        para = para.strip()
        if not para:
            continue
        
        # シーン区切り（====）はスキップ
        if '=' * 10 in para:
            # ====と【シーン:】が同じ段落の場合、【シーン:】だけ抽出
            lines = para.split('\n')
            for line in lines:
                line = line.strip()
                if line.startswith('【シーン:') and '】' in line:
                    heading = line.replace('【シーン:', '').replace('】', '').strip()
                    pages_html.append(f'''    <div class="page">
        <h2>{heading}</h2>
    </div>''')
            continue
        
        # 画像マーカーとテキストが混在している段落を分離
        if ('[画像]' in para or '[背景]' in para or '【--background--】' in para or '【--imagetween--】' in para) and 'https://' in para:
            lines = para.split('\n')
            non_image_lines = []
            
            for line in lines:
                line_stripped = line.strip()
                # 画像行を検出
                if ('[画像]' in line or '[背景]' in line or '【--background--】' in line or '【--imagetween--】' in line) and 'https://' in line:
                    import re as re_module
                    url_match = re_module.search(r'https://[^\s\)\]]+', line)
                    if url_match:
                        url = url_match.group(0)
                        is_background = '[背景]' in line or '【--background--】' in line
                        alt_text = "背景" if is_background else "イラスト"
                        
                        pages_html.append(f'''    <div class="page">
        <img class="illustration" src="{url}" alt="{alt_text}">
    </div>''')
                else:
                    # 画像行でない場合は保存
                    if line_stripped:
                        non_image_lines.append(line_stripped)
            
            # 画像行以外のテキストがあれば処理
            if non_image_lines:
                para = '\n'.join(non_image_lines)
                # 後続の処理のためにparaを更新
            else:
                continue  # 画像のみの段落なので次へ
        
        # 見出しの検出（【シーン: xxx】形式）
        if para.startswith('【シーン:') and '】' in para:
            heading = para.replace('【シーン:', '').replace('】', '').strip()
            heading = heading.replace('\n', '<br>')
            pages_html.append(f'''    <div class="page">
        <h2>{heading}</h2>
    </div>''')
            continue
        
        # 分岐セクションの検出と分割処理
        if '【ドクターの選択肢】' in para or '【分岐:' in para:
            # 分岐セクションを【分岐:】ごとに分割
            branch_pages = _split_branch_section(para)
            pages_html.extend(branch_pages)
            continue
        
        # 通常のテキスト
        # 【話者名】を上に表示する形式に変換
        lines = para.split('\n')
        formatted_lines = []
        for line in lines:
            # 【話者名】セリフ の形式を検出
            if line.strip().startswith('【') and '】' in line:
                # 話者名とセリフを分離
                speaker_end = line.find('】')
                speaker = line[:speaker_end + 1]  # 【話者名】
                dialogue = line[speaker_end + 1:]  # セリフ部分
                
                # 話者名を<span>で囲んで改行を挿入
                formatted_lines.append(f'<span class="speaker">{speaker}</span><br>{dialogue}')
            else:
                formatted_lines.append(line)
        
        formatted = '<br>\n            '.join(formatted_lines)
        formatted = re.sub(r'(\d{1,2})(?=[年月日時分秒cc])', r'<span class="tcy">\1</span>', formatted)
        
        pages_html.append(f'''    <div class="page text">
        <p>
            {formatted}
        </p>
    </div>''')
    
    # HTMLを生成
    final_html = html_template.replace('{pages}', '\n\n'.join(pages_html))
    final_html = final_html.replace('{title}', title)
    
    # ファイルに保存
    output_dir = Path('output')
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / output_file
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(final_html)
    
    return output_path


def _split_branch_section(content):
    """分岐セクションを適切に分割して処理
    
    処理の流れ:
    1. 最初に全選択肢を表示するページ
    2. 各選択肢番号ごとに「選択肢 + 回答」のページを作成
    """
    pages = []
    lines = content.split('\n')
    
    # 全選択肢を収集（最初の選択肢セクションのみ）
    all_choices = []
    branches = {}  # {branch_id: {'choice': str, 'lines': [...]}}
    current_branch_id = None
    current_branch_choice = None
    current_branch_lines = []
    found_first_branch = False  # 最初の分岐マーカーを見つけたかどうか
    
    for line in lines:
        line_stripped = line.strip()
        
        # 【ドクターの選択肢】ヘッダーはスキップ
        if line_stripped.startswith('【ドクターの選択肢】'):
            continue
        
        # 選択肢を収集（最初の分岐前のみ）
        if (line_stripped.startswith('選択肢') or line_stripped.startswith('選択:')) and not found_first_branch:
            all_choices.append(line_stripped)
            continue
        
        # 分岐マーカーを検出
        if line_stripped.startswith('【分岐:'):
            found_first_branch = True
            
            # 前の分岐を保存
            if current_branch_id is not None:
                if current_branch_id not in branches:
                    branches[current_branch_id] = {'choice': current_branch_choice, 'lines': []}
                branches[current_branch_id]['lines'].extend(current_branch_lines)
            
            # 新しい分岐を開始
            # 【分岐: >Options_1】または 【分岐: >Options_1&2&3】から番号を抽出
            import re as re_module
            # 複数選択肢の場合（>Options_1&2&3）
            match = re_module.search(r'>Options_([0-9&]+)', line_stripped)
            if match:
                option_str = match.group(1)
                if '&' in option_str:
                    # 複数選択肢が同じ結果になる場合は、特別なIDを使用
                    current_branch_id = f"combined_{option_str}"
                else:
                    # 単一選択肢
                    current_branch_id = int(option_str)
                current_branch_choice = None  # 次の行で設定される
                current_branch_lines = []
            elif 'End of Options' in line_stripped:
                # 【分岐: End of Options】の場合は特別扱い
                current_branch_id = 'end'
                current_branch_choice = None
                current_branch_lines = []
            continue
        
        # 分岐内の処理
        if current_branch_id is not None:
            # 分岐直後の選択肢表記を保存（表示用としてchoiceに保存し、linesにも含める）
            if (line_stripped.startswith('選択肢') or line_stripped.startswith('選択:')) and current_branch_choice is None:
                current_branch_choice = line_stripped
                # 選択肢もlinesに含める（表示するため）
                current_branch_lines.append(line_stripped)
                continue
            
            # その他の行を保存
            if line_stripped:
                current_branch_lines.append(line_stripped)
    
    # 最後の分岐を保存
    if current_branch_id is not None:
        if current_branch_id not in branches:
            branches[current_branch_id] = {'choice': current_branch_choice, 'lines': []}
        branches[current_branch_id]['lines'].extend(current_branch_lines)
    
    # 全選択肢を表示するページを作成（最初に1回だけ）
    unique_choices = []
    seen = set()
    for choice in all_choices:
        if choice not in seen:
            unique_choices.append(choice)
            seen.add(choice)
    
    if unique_choices:
        pages.append(_create_choices_page(unique_choices))
    
    # 各分岐の処理
    # 1. 単一選択肢（数値キー）: 選択肢 + 回答のページを作成
    # 2. 複数選択肢（combined_キー）: 選択肢表示なしで回答のみ表示
    numeric_branches = {k: v for k, v in branches.items() if isinstance(k, int)}
    combined_branches = {k: v for k, v in branches.items() if isinstance(k, str) and k.startswith('combined_')}
    
    # 単一選択肢の処理
    for i in sorted(numeric_branches.keys()):
        branch_data = numeric_branches[i]
        # 選択肢 + 回答のページを作成
        pages.append(_create_choice_with_response_page(branch_data['choice'], branch_data['lines']))
    
    # 複数選択肢が同じ結果になる場合（combined）の処理
    # 選択肢は既に全選択肢ページで表示済みなので、回答のみ表示
    for branch_id in sorted(combined_branches.keys()):
        branch_data = combined_branches[branch_id]
        # branch_idから番号部分を抽出 (combined_1&2&3 -> 1&2&3)
        option_nums = branch_id.replace('combined_', '')
        pages.append(_create_combined_branch_page(branch_data['lines'], option_nums))
    
    # 'end' 分岐がある場合は、選択肢なしで表示
    if 'end' in branches and branches['end']:
        pages.append(_create_end_branch_page(branches['end']['lines']))
    
    return pages


def _create_choices_page(choices):
    """全選択肢を表示するページを生成"""
    formatted_choices = []
    for choice in choices:
        formatted_choices.append(f'<span class="choice-text">{choice}</span>')
    
    content_html = '<br>\n            '.join(formatted_choices)
    
    return f'''    <div class="page branch text">
        <p>
            {content_html}
        </p>
    </div>'''


def _create_choice_with_response_page(choice_text, response_lines):
    """選択肢 + 回答を表示するページを生成"""
    formatted_lines = []
    
    # 回答を追加（選択肢もresponse_linesに含まれている）
    for line in response_lines:
        if not line:
            continue
        
        # 選択肢行の処理
        if line.startswith('選択肢') or line.startswith('選択:'):
            formatted_lines.append(f'<span class="choice-text">{line}</span>')
        
        # 話者名付きセリフ
        elif line.startswith('【') and '】' in line:
            speaker_end = line.find('】')
            speaker = line[:speaker_end + 1]
            dialogue = line[speaker_end + 1:]
            
            formatted = f'<span class="speaker">{speaker}</span><br>{dialogue}'
            formatted = re.sub(r'(\d{1,2})(?=[年月日時分秒cc])', r'<span class="tcy">\1</span>', formatted)
            formatted_lines.append(formatted)
        
        # その他のテキスト
        else:
            formatted = re.sub(r'(\d{1,2})(?=[年月日時分秒cc])', r'<span class="tcy">\1</span>', line)
            formatted_lines.append(formatted)
    
    content_html = '<br>\n            '.join(formatted_lines)
    
    return f'''    <div class="page branch text">
        <p>
            {content_html}
        </p>
    </div>'''


def _create_end_branch_page(response_lines):
    """選択後の共通セリフを表示するページを生成（選択肢なし）"""
    formatted_lines = []
    
    # マーカーを最初に表示
    formatted_lines.append('<span class="branch-marker">【分岐: End of Options】</span>')
    
    for line in response_lines:
        # 話者名の処理
        if line.startswith('【') and '】' in line:
            # 話者名とセリフを分割
            bracket_end = line.find('】')
            speaker = line[1:bracket_end]
            dialogue = line[bracket_end+1:]
            formatted = f'<span class="speaker">{speaker}</span><br>{dialogue}'
            formatted = re.sub(r'(\d{1,2})(?=[年月日時分秒cc])', r'<span class="tcy">\1</span>', formatted)
            formatted_lines.append(formatted)
        # 選択肢はスキップ
        elif line.startswith('選択肢') or line.startswith('選択:'):
            continue
        # その他のテキスト
        else:
            formatted = re.sub(r'(\d{1,2})(?=[年月日時分秒cc])', r'<span class="tcy">\1</span>', line)
            formatted_lines.append(formatted)
    
    content_html = '<br>\n            '.join(formatted_lines)
    
    return f'''    <div class="page branch text">
        <p>
            {content_html}
        </p>
    </div>'''


def _create_combined_branch_page(response_lines, option_nums):
    """複数選択肢が同じ結果になる場合のページを生成（選択肢表示なし、回答のみ）"""
    formatted_lines = []
    
    # マーカーを最初に表示
    formatted_lines.append(f'<span class="branch-marker">【分岐: >Options_{option_nums}】</span>')
    
    for line in response_lines:
        # 話者名の処理
        if line.startswith('【') and '】' in line:
            # 話者名とセリフを分割
            bracket_end = line.find('】')
            speaker = line[1:bracket_end]
            dialogue = line[bracket_end+1:]
            formatted = f'<span class="speaker">{speaker}</span><br>{dialogue}'
            formatted = re.sub(r'(\d{1,2})(?=[年月日時分秒cc])', r'<span class="tcy">\1</span>', formatted)
            formatted_lines.append(formatted)
        # 選択肢はスキップ（既に全選択肢ページで表示済み）
        elif line.startswith('選択肢') or line.startswith('選択:'):
            continue
        # その他のテキスト
        else:
            formatted = re.sub(r'(\d{1,2})(?=[年月日時分秒cc])', r'<span class="tcy">\1</span>', line)
            formatted_lines.append(formatted)
    
    content_html = '<br>\n            '.join(formatted_lines)
    
    return f'''    <div class="page branch text">
        <p>
            {content_html}
        </p>
    </div>'''


def main():
    import sys
    
    # コマンドライン引数チェック
    skip_ai = '--no-ai' in sys.argv or '--direct' in sys.argv
    
    print("\n" + "=" * 80)
    if skip_ai:
        print("Excel → HTML 直接変換モード (AI変換スキップ)")
    else:
        print("Excel → 小説風HTML 自動変換ツール")
    print("=" * 80 + "\n")
    
    # Excelファイルを検出
    excel_file = DEFAULT_EXCEL_FILE
    if not Path(excel_file).exists():
        # デフォルトがない場合、main_*.xlsxを探す
        excel_files = list(Path('.').glob('main_*.xlsx'))
        if excel_files:
            excel_file = str(excel_files[0])
            print(f"Excelファイル検出: {excel_file}\n")
        else:
            print(f"エラー: {DEFAULT_EXCEL_FILE} が見つかりません。")
            return
    
    # タイトルを抽出
    title = extract_title_from_filename(excel_file)
    print(f"タイトル: {title}\n")
    
    # ステップ1: Excelからデータを抽出（ai_input.txtがない場合のみ）
    input_file_path = Path('output/ai_input.txt')
    if not input_file_path.exists():
        print("ai_input.txtが見つかりません。Excelからデータを抽出します...\n")
        dialogues = extract_all_dialogues(excel_file)
        if dialogues is None:
            return  # openpyxlがない場合は終了
        
        # AI用の入力ファイルを保存
        input_file = save_to_file(dialogues, 'ai_input.txt')
        print(f"\n✓ AIに送信するデータを保存しました: {input_file}")
        print(f"  総文字数: {len(dialogues):,} 文字\n")
        
        # --no-ai オプションの場合は直接HTML生成
        if skip_ai:
            print("=" * 80)
            print("AI変換をスキップして直接HTML生成します")
            print("=" * 80 + "\n")
            output_filename = f'{title}.html'
            html_file = create_html(dialogues, output_file=output_filename, title=title)
            print(f"✓ HTMLファイルを生成しました: {html_file}")
            print(f"\nブラウザで開いて確認してください!")
            return
        
        # プロンプトの例を表示
        print("=" * 80)
        print("【AIへのプロンプト例は ai_prompt.txt を参照してください】")
        print("=" * 80)
        
        print("\n次のステップ:")
        print(f"1. {input_file} をテキストエディタで開く")
        print("2. ai_prompt.txt の内容と一緒にAIに送信")
        print("3. AIの出力を output/novel_output.txt に保存")
        print("4. このスクリプトを再実行")
        print("\nまたは、AI変換をスキップする場合:")
        print("  python simple_converter.py --no-ai\n")
    else:
        print(f"✓ {input_file_path} が既に存在します。")
        print("  Excelからの再抽出をスキップします。\n")
    
    # ステップ2: AI変換後のテキストがあればHTMLを生成
    # タイトル名に基づくファイルを優先、なければnovel_output.txtを使用
    novel_file_with_title = Path(f'output/novel_output_{title}.txt')
    novel_file_default = Path('output/novel_output.txt')
    
    if novel_file_with_title.exists():
        novel_file = novel_file_with_title
    elif novel_file_default.exists():
        novel_file = novel_file_default
    else:
        novel_file = None
    
    if novel_file:
        # 空ファイルチェック
        if novel_file.stat().st_size == 0:
            print(f"⚠ {novel_file.name} は空です。AIの出力を貼り付けてください。")
            return
        
        print("=" * 80)
        print(f"小説テキストが見つかりました: {novel_file.name}")
        print("=" * 80 + "\n")
        
        with open(novel_file, 'r', encoding='utf-8') as f:
            novel_text = f.read()
        
        # 出力ファイル名をタイトルに基づいて決定
        output_filename = f'{title}.html'
        html_file = create_html(novel_text, output_file=output_filename, title=title)
        print(f"✓ HTMLファイルを生成しました: {html_file}")
        print(f"\nブラウザで開いて確認してください!")
    else:
        print(f"⚠ novel_output_{title}.txt または novel_output.txt がありません。")
        print(f"1. output/ai_input.txt をAIに送信")
        print(f"2. AIの出力を output/novel_output_{title}.txt に保存")
        print(f"3. このスクリプトを再実行")

if __name__ == '__main__':
    main()

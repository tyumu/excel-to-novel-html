import openpyxl

# Excelファイルを開く
wb = openpyxl.load_workbook('main_0_暗黒時代・上.xlsx')
sheet = wb.active

print("=== 最初の20行のデータ（全列）===\n")
for i in range(1, min(21, sheet.max_row + 1)):
    col1 = sheet.cell(row=i, column=1).value
    col2 = sheet.cell(row=i, column=2).value
    col3 = sheet.cell(row=i, column=3).value
    
    if col1 or col2 or col3:
        print(f"行{i:3d}: A={col1} | B={col2} | C={col3}")

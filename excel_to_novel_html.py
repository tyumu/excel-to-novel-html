"""
Excelの会話データを小説風HTMLに自動変換するスクリプト

使い方:
1. このスクリプトを実行すると、Excelデータが整形されたテキストファイルが生成されます
2. 生成されたテキストをAIに送信して小説風に変換してもらいま        # 完全なHTMLを組み立て
        html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.title}</title>
    {style_section}
</head>

<body>
{pages_html}
</body>
</html>"""ovel_output.txt に保存します
4. 再度このスクリプトを実行すると、HTMLファイルが生成されます
"""

import openpyxl
import os
import json
import re
from pathlib import Path

class ExcelToNovelConverter:
    def __init__(self, excel_file, template_html='新規 テキスト ドキュメント (5).html'):
        self.excel_file = excel_file
        self.template_html = template_html
        self.output_dir = Path('output')
        self.output_dir.mkdir(exist_ok=True)
        self.title = self._extract_title_from_filename(excel_file)
    
    def _extract_title_from_filename(self, filename):
        """Excelファイル名からタイトルを抽出 (例: main_0_暗黒時代・上.xlsx → 暗黒時代・上)"""
        basename = os.path.basename(filename)
        match = re.match(r'main_\d+_(.+)\.xlsx', basename)
        if match:
            return match.group(1)
        # マッチしない場合は拡張子を除いたファイル名を返す
        return os.path.splitext(basename)[0]
        
    def extract_dialogues_from_excel(self):
        """Excelから会話データを抽出"""
        wb = openpyxl.load_workbook(self.excel_file)
        
        all_sheets_data = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            dialogues = []
            
            for row in range(1, sheet.max_row + 1):
                col2 = sheet.cell(row=row, column=2).value  # 指示列
                col3 = sheet.cell(row=row, column=3).value  # テキスト/URL列
                
                if col3:
                    if col2 == '--image--':
                        dialogues.append({
                            'type': 'image',
                            'url': col3
                        })
                    elif col2 == '----':
                        # 区切り線は無視
                        pass
                    else:
                        # 通常のセリフ
                        dialogues.append({
                            'type': 'text',
                            'content': col3.strip()
                        })
            
            all_sheets_data[sheet_name] = dialogues
        
        return all_sheets_data
    
    def save_for_ai_processing(self, sheets_data, output_file='ai_input.txt'):
        """AIに送信するための整形されたテキストを生成"""
        output_path = self.output_dir / output_file
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("以下のゲームの会話データを、縦書き小説風の文章に変換してください。\n")
            f.write("=" * 80 + "\n\n")
            
            for sheet_name, dialogues in sheets_data.items():
                f.write(f"\n{'=' * 60}\n")
                f.write(f"【シーン: {sheet_name}】\n")
                f.write(f"{'=' * 60}\n\n")
                
                for i, item in enumerate(dialogues, 1):
                    if item['type'] == 'image':
                        f.write(f"[画像 {i}]: {item['url']}\n\n")
                    else:
                        f.write(f"{i}. {item['content']}\n\n")
                
                f.write("\n" + "-" * 60 + "\n")
        
        print(f"✓ AIに送信するデータを保存しました: {output_path}")
        print(f"\n次のステップ:")
        print(f"1. {output_path} の内容をAIに送信して小説風に変換してもらってください")
        print(f"2. AIの出力を {self.output_dir / 'novel_output.txt'} に保存してください")
        print(f"3. このスクリプトを再実行すると、HTMLが生成されます")
        
        return output_path
    
    def generate_html_from_novel(self, novel_file='novel_output.txt'):
        """AIで変換された小説テキストからHTMLを生成"""
        novel_path = self.output_dir / novel_file
        
        if not novel_path.exists():
            print(f"エラー: {novel_path} が見つかりません")
            print("まず、AIで変換した小説テキストを保存してください")
            return None
        
        # テンプレートHTMLを読み込み
        with open(self.template_html, 'r', encoding='utf-8') as f:
            template = f.read()
        
        # テンプレートからスタイル部分を抽出
        style_start = template.find('<style>')
        style_end = template.find('</style>') + len('</style>')
        style_section = template[style_start:style_end]
        
        # ヘッダー部分を抽出
        head_end = template.find('</head>')
        header = template[:head_end]
        
        # 小説テキストを読み込み
        with open(novel_path, 'r', encoding='utf-8') as f:
            novel_content = f.read()
        
        # HTMLページを生成
        pages_html = self._parse_novel_to_pages(novel_content)
        
        # 完全なHTMLを組み立て
        html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.title}</title>
    {style_section}
</head>

<body>
{pages_html}
</body>
</html>"""
        
        # HTMLファイルを保存
        output_html = self.output_dir / 'novel_generated.html'
        with open(output_html, 'w', encoding='utf-8') as f:
            f.write(html)
        
        print(f"✓ HTMLファイルを生成しました: {output_html}")
        return output_html
    
    def _parse_novel_to_pages(self, content):
        """小説テキストをHTMLページに変換"""
        pages = []
        lines = content.split('\n')
        
        current_page_lines = []
        max_lines_per_page = 10  # 1ページあたりの最大行数
        
        for line in lines:
            line = line.strip()
            
            if not line:
                continue
            
            # 画像URLの検出
            if line.startswith('[画像') or 'https://' in line:
                if current_page_lines:
                    pages.append(self._create_text_page('\n'.join(current_page_lines)))
                    current_page_lines = []
                
                # 画像ページを追加
                url_start = line.find('https://')
                if url_start != -1:
                    url = line[url_start:].strip()
                    pages.append(self._create_image_page(url))
                continue
            
            # 見出しの検出
            if line.startswith('【') and line.endswith('】'):
                if current_page_lines:
                    pages.append(self._create_text_page('\n'.join(current_page_lines)))
                    current_page_lines = []
                
                heading = line[1:-1]  # 【】を除去
                pages.append(self._create_heading_page(heading))
                continue
            
            # 通常のテキスト
            current_page_lines.append(line)
            
            # ページが満杯になったら次のページへ
            if len(current_page_lines) >= max_lines_per_page:
                pages.append(self._create_text_page('\n'.join(current_page_lines)))
                current_page_lines = []
        
        # 残りのテキストをページに追加
        if current_page_lines:
            pages.append(self._create_text_page('\n'.join(current_page_lines)))
        
        return '\n'.join(pages)
    
    def _create_text_page(self, text):
        """テキストページのHTMLを生成"""
        # 改行を<br>タグに変換
        formatted_text = text.replace('\n', '<br>\n            ')
        return f"""    <div class="page text">
        <p>
            {formatted_text}
        </p>
    </div>
"""
    
    def _create_heading_page(self, heading):
        """見出しページのHTMLを生成"""
        # 改行を<br>タグに変換
        formatted_heading = heading.replace('\n', '<br>')
        return f"""    <div class="page">
        <h2>{formatted_heading}</h2>
    </div>
"""
    
    def _create_image_page(self, url):
        """画像ページのHTMLを生成"""
        return f"""    <div class="page">
        <img class="illustration" src="{url}" alt="イラスト">
    </div>
"""


def main():
    converter = ExcelToNovelConverter('main_0_暗黒時代・上.xlsx')
    
    # ステップ1: Excelからデータを抽出してAI用のテキストを生成
    print("=" * 80)
    print("Excelデータを抽出中...")
    print("=" * 80)
    
    sheets_data = converter.extract_dialogues_from_excel()
    
    print(f"\n抽出したシート数: {len(sheets_data)}")
    for sheet_name, dialogues in sheets_data.items():
        print(f"  - {sheet_name}: {len(dialogues)}項目")
    
    print("\n" + "=" * 80)
    converter.save_for_ai_processing(sheets_data)
    print("=" * 80)
    
    # ステップ2: AI変換後のテキストがあればHTMLを生成
    novel_file = converter.output_dir / 'novel_output.txt'
    if novel_file.exists():
        print("\n" + "=" * 80)
        print("小説テキストが見つかりました。HTMLを生成します...")
        print("=" * 80)
        converter.generate_html_from_novel()


if __name__ == '__main__':
    main()

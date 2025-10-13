import openpyxl

# Excelファイルを開く
wb = openpyxl.load_workbook('main_0_暗黒時代・上.xlsx')

# 各シートで話者情報(B列)があるか確認
for sheet_name in wb.sheetnames[:5]:  # 最初の5シートを確認
    sheet = wb[sheet_name]
    print(f"\n{'=' * 60}")
    print(f"シート: {sheet_name}")
    print(f"{'=' * 60}")
    
    has_speaker = False
    for i in range(1, min(30, sheet.max_row + 1)):
        col2 = sheet.cell(row=i, column=2).value
        col3 = sheet.cell(row=i, column=3).value
        
        # B列に話者情報がある場合（----や--image--以外）
        if col2 and col2 not in ['----', '--image--', '--imagetween--']:
            has_speaker = True
            print(f"行{i:3d}: 話者=[{col2}] セリフ=[{col3}]")
    
    if not has_speaker:
        print("→ このシートには話者情報がありません")
        # サンプルを表示
        print("\nサンプル:")
        for i in range(1, min(6, sheet.max_row + 1)):
            col2 = sheet.cell(row=i, column=2).value
            col3 = sheet.cell(row=i, column=3).value
            if col3:
                print(f"  行{i}: B={col2} | C={col3}")

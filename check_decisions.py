import openpyxl

# Excelファイルを開く
wb = openpyxl.load_workbook('main_0_暗黒時代・上.xlsx')
sheet = wb['level_main_00-01_end']

print("=== Decision分岐の例（level_main_00-01_end シート）===\n")
for i in range(20, 35):
    col1 = sheet.cell(row=i, column=1).value
    col2 = sheet.cell(row=i, column=2).value
    col3 = sheet.cell(row=i, column=3).value
    
    if col1 or col2 or col3:
        print(f"行{i:3d}: A={col1:4} | B={col2:20} | C={col3}")

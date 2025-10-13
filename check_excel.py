import openpyxl

# Excelファイルを開く
wb = openpyxl.load_workbook('main_0_暗黒時代・上.xlsx')
print(f"シート名: {wb.sheetnames}")

# 最初のシートを取得
sheet = wb.active
print(f"\nアクティブシート: {sheet.title}")
print(f"最大行数: {sheet.max_row}")
print(f"最大列数: {sheet.max_column}")

# 最初の10行を表示
print("\n=== 最初の10行のサンプル ===")
for i in range(1, min(11, sheet.max_row + 1)):
    row_data = []
    for j in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=i, column=j).value
        if cell_value:
            row_data.append(f"列{j}: {cell_value}")
    if row_data:
        print(f"行{i}: {' | '.join(row_data)}")

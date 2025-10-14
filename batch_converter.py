"""
複数のExcelファイルを一括処理して個別のHTMLを生成

フォルダ内の main_*.xlsx ファイルをすべて処理します。
各Excelファイルに対して:
1. ai_input_[ファイル名].txt を生成
2. AIで変換した novel_output_[ファイル名].txt から
3. [ファイル名].html を生成

使い方: python batch_converter.py
"""

import re
from pathlib import Path
from simple_converter import extract_all_dialogues, save_to_file, create_html

def get_excel_files():
    """フォルダ内のすべてのmain_*.xlsxファイルを取得"""
    current_dir = Path('.')
    excel_files = list(current_dir.glob('main_*.xlsx'))
    return sorted(excel_files)

def extract_title_from_filename(filename):
    """ファイル名からタイトルを抽出
    例: main_0_暗黒時代・上.xlsx → main_0_暗黒時代・上
    """
    # .xlsx拡張子のみを削除
    match = re.match(r'(main_\d+_.+)\.xlsx', filename)
    if match:
        return match.group(1)
    return filename.replace('.xlsx', '')

def extract_display_title(filename):
    """ファイル名から表示用タイトルを抽出(main_数字_を除く)
    例: main_0_暗黒時代・上.xlsx → 暗黒時代・上
    """
    match = re.match(r'main_\d+_(.+)\.xlsx', filename)
    if match:
        return match.group(1)
    return filename.replace('.xlsx', '')

def process_excel_file(excel_path, skip_ai=False):
    """1つのExcelファイルを処理
    
    Args:
        excel_path: Excelファイルのパス
        skip_ai: Trueの場合、AI変換をスキップして直接HTML生成
    """
    print("\n" + "=" * 80)
    print(f"処理中: {excel_path.name}")
    print("=" * 80)
    
    # タイトル抽出
    title = extract_title_from_filename(excel_path.name)
    display_title = extract_display_title(excel_path.name)
    print(f"タイトル: {title}")
    print(f"表示タイトル: {display_title}")
    
    # 出力ファイル名を決定
    ai_input_file = f'ai_input_{title}.txt'
    novel_output_file = f'novel_output_{title}.txt'
    html_file = f'{title}.html'
    
    # ai_input ファイルが存在するかチェック
    ai_input_path = Path('output') / ai_input_file
    
    if not ai_input_path.exists():
        print(f"\n{ai_input_file} が見つかりません。Excelからデータを抽出します...\n")
        
        # Excelファイル名を一時的に保存
        import simple_converter
        original_excel = getattr(simple_converter, 'EXCEL_FILE', None)
        
        # extract_all_dialogues を直接呼び出す代わりに、Excelを読み込む
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            
            all_text = []
            decision_count = 0
            
            # config から設定を読み込み
            try:
                from config import DOCTOR_NAME
            except ImportError:
                DOCTOR_NAME = "ドクター"
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                all_text.append(f"\n\n{'=' * 60}")
                all_text.append(f"【シーン: {sheet_name}】")
                all_text.append(f"{'=' * 60}\n")
                
                in_decision = False
                current_options = []
                choices_already_displayed = False  # 選択肢を一度表示したかどうか
                
                for row in range(1, sheet.max_row + 1):
                    col2 = sheet.cell(row=row, column=2).value
                    col3 = sheet.cell(row=row, column=3).value
                    
                    if not col3:
                        continue
                    
                    # 分岐システムの処理
                    if col2 == '--Decision--':
                        in_decision = True
                        decision_count += 1
                        current_options = []
                        choices_already_displayed = False  # リセット
                        continue
                    
                    if col2 == '--Decision End--':
                        in_decision = False
                        # 全ての選択肢を最初の1回だけ表示
                        if current_options and not choices_already_displayed:
                            all_text.append('\n【ドクターの選択肢】')
                            all_text.extend(current_options)
                            all_text.append('')
                            choices_already_displayed = True
                        continue
                    
                    if in_decision and col2 and col2.startswith('Option_'):
                        option_num = col2.replace('Option_', '')
                        current_options.append(f"  選択肢{option_num}: {col3}")
                        continue
                    
                    if col2 == '--Branch--':
                        branch_info = col3.strip() if col3 else ''
                        all_text.append(f'\n【分岐: {branch_info}】')
                        
                        # >Options_X または >Options_X&Y&Z から番号を抽出
                        import re
                        # >Options_1&2&3 のような複数選択肢にも対応
                        match = re.search(r'>Options_([0-9&]+)', branch_info)
                        if match and current_options:
                            option_nums_str = match.group(1)
                            # &で分割して複数の番号を取得
                            option_nums = [int(n) for n in option_nums_str.split('&')]
                            # 各選択肢を表示
                            for option_num in option_nums:
                                if 0 < option_num <= len(current_options):
                                    all_text.append(current_options[option_num - 1])
                        continue
                    
                    # 画像
                    if col2 == '--imagetween--':
                        all_text.append(f'\n[画像]: {col3}')
                        continue
                    
                    if col2 == '--image--':
                        all_text.append(f'\n[背景]: {col3}')
                        continue
                    
                    # 話者とセリフ
                    if col2 and col2 not in ['----', '--imagetween--', '--Decision--', '--Decision End--', '--Branch--']:
                        speaker = col2.strip()
                        dialogue = col3.strip() if col3 else ''
                        
                        # {@nickname} を DOCTOR_NAME に置換
                        dialogue = dialogue.replace('{@nickname}', DOCTOR_NAME)
                        
                        all_text.append(f'【{speaker}】{dialogue}')
            
            dialogues = '\n'.join(all_text)
            
            print(f"\n検出した分岐数: {decision_count}")
            print(f"総文字数: {len(dialogues):,} 文字")
            
            # 保存
            output_path = save_to_file(dialogues, ai_input_file)
            print(f"\n✓ AIに送信するデータを保存しました: {output_path}")
            
            # --no-ai オプションの場合は直接HTML生成
            if skip_ai:
                print(f"\n✓ AI変換をスキップして直接HTML生成します...")
                html_path = create_html(dialogues, output_file=html_file, title=display_title)
                print(f"✓ HTMLファイルを生成しました: {html_path}")
                return True
            
            # 空の novel_output ファイルも生成
            novel_output_path = Path('output') / novel_output_file
            if not novel_output_path.exists():
                novel_output_path.touch()
                print(f"✓ 空の {novel_output_file} を作成しました（AIの出力をここに貼り付けてください）")
            
        except Exception as e:
            print(f"エラーが発生しました: {e}")
            return False
    else:
        print(f"✓ {ai_input_file} が既に存在します。")
        
        # --no-ai オプションの場合は ai_input から直接HTML生成
        if skip_ai:
            ai_input_path = Path('output') / ai_input_file
            if ai_input_path.exists() and ai_input_path.stat().st_size > 0:
                print(f"\n✓ AI変換をスキップして直接HTML生成します...")
                with open(ai_input_path, 'r', encoding='utf-8') as f:
                    dialogues = f.read()
                html_path = create_html(dialogues, output_file=html_file, title=display_title)
                print(f"✓ HTMLファイルを生成しました: {html_path}")
                return True
        
        # 空の novel_output ファイルも生成（存在しない場合のみ）
        novel_output_path_temp = Path('output') / novel_output_file
        if not novel_output_path_temp.exists():
            novel_output_path_temp.touch()
            print(f"✓ 空の {novel_output_file} を作成しました（AIの出力をここに貼り付けてください）")
    
    # novel_output ファイルが存在するかチェック
    novel_output_path = Path('output') / novel_output_file
    
    if novel_output_path.exists():
        # ファイルサイズをチェック（空ファイルはスキップ）
        if novel_output_path.stat().st_size == 0:
            print(f"\n⚠ {novel_output_file} は空です。AIの出力を貼り付けてください。")
            return False
        
        print(f"\n✓ {novel_output_file} が見つかりました。HTMLを生成します...")
        
        with open(novel_output_path, 'r', encoding='utf-8') as f:
            novel_text = f.read()
        
        html_path = create_html(novel_text, output_file=html_file, title=display_title)
        print(f"✓ HTMLファイルを生成しました: {html_path}")
        return True
    else:
        print(f"\n⚠ {novel_output_file} がまだありません。")
        print(f"1. output/{ai_input_file} をAIに送信")
        print(f"2. AIの出力を output/{novel_output_file} として保存")
        print(f"3. このスクリプトを再実行")
        return False

def main():
    import sys
    
    # コマンドライン引数チェック
    skip_ai = '--no-ai' in sys.argv or '--direct' in sys.argv
    
    print("\n" + "=" * 80)
    if skip_ai:
        print("Excel → HTML 一括直接変換ツール (AI変換スキップ)")
    else:
        print("Excel → HTML 一括変換ツール")
    print("=" * 80 + "\n")
    
    # Excelファイルを取得
    excel_files = get_excel_files()
    
    if not excel_files:
        print("エラー: main_*.xlsx ファイルが見つかりません。")
        print("このスクリプトと同じフォルダに Excel ファイルを配置してください。")
        return
    
    print(f"見つかったExcelファイル: {len(excel_files)}個\n")
    for f in excel_files:
        print(f"  - {f.name}")
    
    print("\n処理を開始します...\n")
    
    # 各ファイルを処理
    results = {}
    for excel_file in excel_files:
        success = process_excel_file(excel_file, skip_ai=skip_ai)
        results[excel_file.name] = success
    
    # 結果サマリー
    print("\n" + "=" * 80)
    print("処理結果サマリー")
    print("=" * 80 + "\n")
    
    completed = sum(1 for v in results.values() if v)
    pending = sum(1 for v in results.values() if not v)
    
    print(f"✓ HTML生成完了: {completed}個")
    print(f"⚠ AI変換待ち: {pending}個\n")
    
    if pending > 0:
        print("AI変換待ちのファイル:")
        for filename, success in results.items():
            if not success:
                title = extract_title_from_filename(filename)
                print(f"  - {title} (output/ai_input_{title}.txt → output/novel_output_{title}.txt)")

if __name__ == '__main__':
    main()
